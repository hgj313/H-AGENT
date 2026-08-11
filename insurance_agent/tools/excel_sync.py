"""Excel 模板同步工具

将智能体提取的增减保人员同步到 Excel 模板文件。
- 减保人员: 从 Excel 中删除（按证件号码匹配）
- 增保人员: 如果 Excel 中不存在则新增，已存在则跳过
- 不改变 Excel 原有字段结构，只填有数据的字段

Excel 模板字段: 姓名 | 证件号码 | 年龄 | 打卡项目 | 所属班组 | 所属公司 | 劳务分类
智能体提取字段: name, id_number, company, birth_date, modification_type
"""

import logging
import os
import shutil
from typing import Optional

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Excel 模板列名 → 列索引（从 1 开始）
# 姓名 | 证件号码 | 年龄 | 打卡项目 | 所属班组 | 所属公司 | 劳务分类
COL_NAME = 1
COL_ID_NUMBER = 2
COL_AGE = 3
COL_COMPANY = 6


def _calc_age(birth_date: str) -> Optional[str]:
    """从出生日期计算年龄（简单计算，不做月日精确判断）"""
    if not birth_date or len(birth_date) < 4:
        return None
    try:
        year = int(birth_date[:4])
        if 1940 <= year <= 2039:
            from datetime import datetime
            return str(datetime.now().year - year)
    except (ValueError, IndexError):
        pass
    return None


def _calc_age_from_id(id_number: str) -> Optional[str]:
    """从身份证号提取年龄"""
    if not id_number or len(id_number) < 14:
        return None
    try:
        year = int(id_number[6:10])
        if 1940 <= year <= 2039:
            from datetime import datetime
            return str(datetime.now().year - year)
    except (ValueError, IndexError):
        pass
    return None


def sync_excel_with_extraction(
    excel_path: str,
    extraction_results: list[dict],
    output_path: Optional[str] = None,
) -> dict:
    """将提取的增减保人员同步到 Excel 模板

    Args:
        excel_path: Excel 模板文件路径
        extraction_results: 智能体提取结果列表，每个元素是一个保单的结果 dict，
                            包含 insured_persons 字段
        output_path: 输出路径，默认覆盖原文件

    Returns:
        统计信息 dict: {
            "added": 新增人数,
            "removed": 删除人数,
            "skipped": 跳过人数（已存在）,
            "total_in_excel": 操作后 Excel 总人数,
        }
    """
    if output_path is None:
        output_path = excel_path

    # 同步前自动备份原文件（仅在覆盖原文件时）
    backup_path = None
    if output_path == excel_path and os.path.exists(excel_path):
        base, ext = os.path.splitext(excel_path)
        backup_path = f"{base}_backup{ext}"
        shutil.copy2(excel_path, backup_path)
        logger.info("已自动备份原文件到 %s", backup_path)

    wb = load_workbook(excel_path)
    ws = wb.active

    # 收集 Excel 中已有的证件号码 → 行号映射
    existing_ids: dict[str, int] = {}  # id_number → row_index
    for row_idx in range(2, ws.max_row + 1):
        id_val = ws.cell(row=row_idx, column=COL_ID_NUMBER).value
        if id_val:
            id_str = str(id_val).strip()
            if id_str:
                existing_ids[id_str] = row_idx

    logger.info("Excel 模板现有 %d 人", len(existing_ids))

    # 从提取结果中收集所有增减保人员
    persons_to_add = []  # 增保人员
    persons_to_remove = []  # 减保人员

    for result in extraction_results:
        if result.get("error"):
            continue
        for person in result.get("insured_persons", []):
            mod_type = person.get("modification_type", "")
            id_number = (person.get("id_number") or "").strip()
            if not id_number:
                continue
            if mod_type == "减保":
                persons_to_remove.append(id_number)
            elif mod_type == "增保":
                persons_to_add.append(person)

    # 执行删除: 减保人员从 Excel 中移除
    removed_count = 0
    # 从后往前删，避免行号变化问题
    rows_to_delete = []
    for id_number in persons_to_remove:
        if id_number in existing_ids:
            rows_to_delete.append(existing_ids[id_number])
            del existing_ids[id_number]
            removed_count += 1

    # 行号从大到小排序，从后往前删
    rows_to_delete.sort(reverse=True)
    for row_idx in rows_to_delete:
        ws.delete_rows(row_idx, 1)

    logger.info("删除减保人员 %d 人", removed_count)

    # 执行新增: 增保人员如果不存在则添加
    added_count = 0
    skipped_count = 0

    # 找到当前最后一行有数据的行号
    current_max_row = ws.max_row
    # 过滤空行，找到真正的最后一行
    while current_max_row > 1:
        if ws.cell(row=current_max_row, column=COL_NAME).value:
            break
        current_max_row -= 1

    for person in persons_to_add:
        id_number = (person.get("id_number") or "").strip()
        if id_number in existing_ids:
            skipped_count += 1
            continue

        # 新增一行
        current_max_row += 1
        ws.cell(row=current_max_row, column=COL_NAME, value=person.get("name", ""))

        # 证件号码存为字符串，避免科学计数法
        id_val = person.get("id_number", "")
        ws.cell(row=current_max_row, column=COL_ID_NUMBER, value=str(id_val) if id_val else "")

        # 年龄: 优先从 birth_date 计算，其次从身份证号
        age = _calc_age(person.get("birth_date", "")) or _calc_age_from_id(id_number)
        if age:
            ws.cell(row=current_max_row, column=COL_AGE, value=age)

        # 所属公司
        ws.cell(row=current_max_row, column=COL_COMPANY, value=person.get("company", ""))

        existing_ids[id_number] = current_max_row
        added_count += 1

    logger.info("新增增保人员 %d 人, 跳过已存在 %d 人", added_count, skipped_count)

    # 保存
    wb.save(output_path)
    logger.info("Excel 已保存到 %s, 总计 %d 人", output_path, len(existing_ids))

    return {
        "added": added_count,
        "removed": removed_count,
        "skipped": skipped_count,
        "total_in_excel": len(existing_ids),
        "backup_path": backup_path,
    }


def sync_excel_from_json(
    excel_path: str,
    json_path: str,
    output_path: Optional[str] = None,
) -> dict:
    """从 extraction_results.json 文件读取提取结果并同步到 Excel

    Args:
        excel_path: Excel 模板文件路径
        json_path: 提取结果 JSON 文件路径
        output_path: 输出路径

    Returns:
        同 sync_excel_with_extraction
    """
    import json

    with open(json_path, "r", encoding="utf-8") as f:
        extraction_results = json.load(f)

    return sync_excel_with_extraction(excel_path, extraction_results, output_path)
