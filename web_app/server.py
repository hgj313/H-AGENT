"""保险单识别 Web 服务

提供前端界面上传 PDF 附件，提取被保人员信息，输出表格文件。
基于现有 insurance_agent 智能体，不修改 Agent 行为，仅做 Web 封装。
"""

import csv
import io
import json
import os
import sys
import tempfile
import traceback
from datetime import datetime
from urllib.parse import quote

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse, FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pydantic import BaseModel

# 确保项目根目录在 path 中
sys.path.insert(0, "C:/insurance-automation")

from dotenv import load_dotenv
load_dotenv("C:/insurance-automation/H-AGENT/.env")

from insurance_agent.infrastructure.parsers import PyMuPDFParser
from insurance_agent.infrastructure import PolicyLibrary
from insurance_agent.infrastructure.llm.factory import create_minimax_llm_from_env
from insurance_agent.agents.invoice_recognition import (
    InvoiceRecognitionCapability,
    create_invoice_recognition_state,
    build_invoice_recognition_graph,
)
from insurance_agent.domain import ExtractionResult, InsuredPerson
from insurance_agent.tools import parse_policy_filename, is_main_policy, is_endorsement

app = FastAPI(title="保险单识别系统", version="1.0.0")

# 静态文件
app.mount("/static", StaticFiles(directory=os.path.join(os.path.dirname(__file__), "static")), name="static")

# 全局状态
_llm_client = None
_policy_library = PolicyLibrary(base_dir="C:/insurance-automation/policy_library")
_latest_results: list[dict] = []  # 最近一次提取结果


def get_llm():
    """获取 LLM 客户端（延迟初始化）"""
    global _llm_client
    if _llm_client is None:
        try:
            _llm_client = create_minimax_llm_from_env()
        except Exception as e:
            print(f"[WARN] LLM 客户端创建失败: {e}")
    return _llm_client


def _fix_filename(filename: str) -> str:
    """修复中文文件名编码问题

    浏览器/ multipart 上传的中文文件名可能被 Python multipart 库
    以 Latin-1 解码（实际是 GBK 或 UTF-8 字节），导致乱码。
    尝试重新编码还原正确中文。
    """
    if not filename:
        return filename
    # 已经是合法中文则直接返回
    if any('\u4e00' <= c <= '\u9fff' for c in filename):
        return filename
    # 尝试 latin-1 → gbk（Windows 浏览器常见）
    try:
        fixed = filename.encode('latin-1').decode('gbk')
        return fixed
    except (UnicodeDecodeError, UnicodeEncodeError):
        pass
    # 尝试 latin-1 → utf-8
    try:
        fixed = filename.encode('latin-1').decode('utf-8')
        return fixed
    except (UnicodeDecodeError, UnicodeEncodeError):
        pass
    return filename


def run_agent(pdf_path: str, llm_client=None, policy_library=None) -> dict:
    """运行 Agent 识别单份保单（与 test_agent.py 逻辑一致）"""
    pdf_parser = PyMuPDFParser()
    capability = InvoiceRecognitionCapability(
        pdf_parser=pdf_parser,
        llm_client=llm_client,
        policy_library=policy_library,
    )
    graph = build_invoice_recognition_graph(capability)
    initial_state = create_invoice_recognition_state(
        user_goal=f"提取 {pdf_path} 中的被保人员清单",
        file_path=pdf_path,
    )
    final_state = graph.invoke(initial_state)
    return final_state


def process_files(file_paths: list[str]) -> list[dict]:
    """批量处理 PDF 文件（先保单后批单）"""
    llm = get_llm()

    # 按保单类型排序
    main_files = [f for f in file_paths if is_main_policy(f)]
    batch_files = [f for f in file_paths if is_endorsement(f)]
    other_files = [f for f in file_paths if not is_main_policy(f) and not is_endorsement(f)]
    sorted_files = main_files + other_files + batch_files

    results = []
    for fpath in sorted_files:
        fname = os.path.basename(fpath)
        try:
            final_state = run_agent(fpath, llm_client=llm, policy_library=_policy_library)
            result_dict = final_state.get("extraction_result") or {
                "file_name": fname,
                "error": final_state.get("error"),
            }
            result_dict["file_path"] = fpath

            fname_info = parse_policy_filename(fname)
            if not result_dict.get("policy_holder"):
                result_dict["policy_holder"] = fname_info.company

            # 注册到保单文件库
            if not result_dict.get("error"):
                _policy_library.register(result_dict)

            results.append(result_dict)
        except Exception as e:
            traceback.print_exc()
            results.append({"file_name": fname, "error": str(e)})

    return results


def results_to_rows(results: list[dict]) -> list[dict]:
    """将提取结果转为扁平化行"""
    all_rows = []
    for r in results:
        if r.get("error"):
            continue
        result = ExtractionResult(
            file_name=r.get("file_name", ""),
            insurance_company=r.get("insurance_company", ""),
            policy_number=r.get("policy_number", ""),
            overall_start_date=r.get("overall_start_date", ""),
            overall_end_date=r.get("overall_end_date", ""),
            insured_persons=[InsuredPerson(**p) for p in r.get("insured_persons", [])],
        )
        all_rows.extend(result.to_csv_rows())
    return all_rows


CSV_FIELDS = [
    "姓名", "证件号码", "证件类型", "出生日期",
    "所属公司", "批改类型",
    "起始时间", "起止时间",
    "岗位名称", "保险公司", "保单号", "来源文件",
]


# ==================== API ====================

@app.get("/api/health")
async def health():
    return {"status": "ok", "llm_available": _llm_client is not None}


@app.post("/api/upload")
async def upload_files(files: list[UploadFile] = File(...)):
    """上传多个 PDF 文件并提取被保人员信息"""
    global _latest_results

    if not files:
        raise HTTPException(status_code=400, detail="未上传文件")

    saved_paths = []
    tmp_dir = tempfile.mkdtemp(prefix="insurance_upload_")

    for f in files:
        raw_name = f.filename or ""
        filename = _fix_filename(raw_name)
        if not filename.lower().endswith(".pdf"):
            continue
        save_path = os.path.join(tmp_dir, filename)
        with open(save_path, "wb") as out:
            content = await f.read()
            out.write(content)
        saved_paths.append(save_path)

    if not saved_paths:
        raise HTTPException(status_code=400, detail="未找到 PDF 文件")

    # 处理文件
    results = process_files(saved_paths)
    _latest_results = results

    # 构建返回数据
    summary = []
    for r in results:
        if r.get("error"):
            summary.append({
                "file_name": r.get("file_name", ""),
                "error": r["error"],
                "persons": [],
            })
            continue
        persons = r.get("insured_persons", [])
        add_count = sum(1 for p in persons if p.get("modification_type") == "增保")
        remove_count = sum(1 for p in persons if p.get("modification_type") == "减保")
        summary.append({
            "file_name": r.get("file_name", ""),
            "insurance_company": r.get("insurance_company", ""),
            "policy_number": r.get("policy_number", ""),
            "overall_start_date": r.get("overall_start_date", ""),
            "overall_end_date": r.get("overall_end_date", ""),
            "persons_count": len(persons),
            "add_count": add_count,
            "remove_count": remove_count,
            "persons": persons,
        })

    total_persons = sum(s.get("persons_count", 0) for s in summary)
    total_add = sum(s.get("add_count", 0) for s in summary)
    total_remove = sum(s.get("remove_count", 0) for s in summary)

    return JSONResponse({
        "success": True,
        "total_files": len(summary),
        "total_persons": total_persons,
        "total_add": total_add,
        "total_remove": total_remove,
        "results": summary,
    })


@app.get("/api/download/csv")
async def download_csv():
    """下载 CSV 表格"""
    if not _latest_results:
        raise HTTPException(status_code=404, detail="无提取结果，请先上传文件")

    rows = results_to_rows(_latest_results)
    if not rows:
        raise HTTPException(status_code=404, detail="无可导出的数据")

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=CSV_FIELDS)
    writer.writeheader()
    writer.writerows(rows)

    filename = f"被保人员清单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    encoded_filename = quote(filename)
    return StreamingResponse(
        iter([output.getvalue().encode("utf-8-sig")]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@app.get("/api/download/xlsx")
async def download_xlsx():
    """下载 Excel 表格"""
    if not _latest_results:
        raise HTTPException(status_code=404, detail="无提取结果，请先上传文件")

    rows = results_to_rows(_latest_results)
    if not rows:
        raise HTTPException(status_code=404, detail="无可导出的数据")

    wb = Workbook()
    ws = wb.active
    ws.title = "被保人员清单"

    # 表头样式
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 写表头
    for col_idx, field in enumerate(CSV_FIELDS, 1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 写数据
    data_font = Font(name="微软雅黑", size=10)
    add_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")  # 浅红
    remove_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # 浅蓝

    for row_idx, row in enumerate(rows, 2):
        for col_idx, field in enumerate(CSV_FIELDS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(field, ""))
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            # 增保/减保行着色
            if field == "批改类型":
                if row.get(field) == "增保":
                    cell.fill = add_fill
                elif row.get(field) == "减保":
                    cell.fill = remove_fill

    # 自动列宽
    col_widths = {
        "姓名": 12, "证件号码": 24, "证件类型": 10, "出生日期": 14,
        "所属公司": 28, "批改类型": 10,
        "起始时间": 14, "起止时间": 14,
        "岗位名称": 16, "保险公司": 20, "保单号": 30, "来源文件": 40,
    }
    for col_idx, field in enumerate(CSV_FIELDS, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = col_widths.get(field, 15)

    # 冻结首行
    ws.freeze_panes = "A2"

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"被保人员清单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@app.get("/api/policy-library")
async def get_policy_library():
    """获取保单文件库状态"""
    records = []
    for r in _policy_library.records:
        records.append({
            "file_name": r.file_name,
            "policy_type": r.policy_type,
            "policy_number": r.policy_number,
            "company": r.company,
            "insurance_company": r.insurance_company,
            "start_date": r.start_date,
            "end_date": r.end_date,
            "persons_count": r.persons_count,
        })
    return {"records": records, "total": len(records)}


@app.get("/")
async def index():
    """返回前端页面"""
    return FileResponse(os.path.join(os.path.dirname(__file__), "static", "index.html"))


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        app,
        host="0.0.0.0",
        port=8765,
        workers=1,
    )
